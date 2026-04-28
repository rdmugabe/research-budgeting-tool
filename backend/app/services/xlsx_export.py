"""xlsx writers for the Final Budget Presentation and PRA workbook.

Both writers consume the ComputedBudget pydantic object produced by the
pricing engine, so exports are 1:1 reproducible from any (frozen) round.

Formatting is functional, not pixel-faithful to the originals. The output is
suitable to send to a sponsor as-is or to paste into a branded template.
"""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.budget import ComputedBudget

# --- Styling helpers ---

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11)
TOTAL_FONT = Font(bold=True, size=11)

THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = "$#,##0.00"
PERCENT = "0.00%"
INTEGER = "#,##0"


def _style_header_row(ws, row_idx: int, last_col: int):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def _autosize(ws, max_widths: dict[int, int]):
    for col, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


# --- Final Budget Presentation ---

def write_final_budget(comp: ComputedBudget, trial_name: str) -> bytes:
    """Produce a single-sheet workbook summarizing per-visit cost and the
    site-fee / pass-through tables for this round.

    Per-visit value is the per-subject-per-visit-completed cost: the sum of
    procedure unit totals (inclusive of OH) for procedures applicable at that
    visit, not multiplied by the visit's completion count.
    """
    # Per-visit per-subject cost = sum of unit_total_with_oh across applicable lines.
    per_subject_by_visit: dict[str, float] = defaultdict(float)
    for line in comp.procedure_lines:
        per_subject_by_visit[line.visit_label] += line.unit_total_with_oh

    wb = Workbook()
    ws = wb.active
    ws.title = "Final Budget Presentation"

    row = 1
    ws.cell(row=row, column=1, value=f"FINAL BUDGET PRESENTATION — {trial_name}").font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    ws.cell(row=row, column=1, value=f"Round {comp.round_number}: {comp.round_label}{'  (FROZEN)' if comp.is_frozen else ''}").font = Font(italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Per-visit table
    headers = ["Visit", "Per Subject Per Visit (incl. OH)", "Completion Count", "Visit Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    visit_section_start = row
    for v in comp.visits:
        ws.cell(row=row, column=1, value=v.visit_label)
        ws.cell(row=row, column=2, value=per_subject_by_visit.get(v.visit_label, 0.0)).number_format = MONEY
        ws.cell(row=row, column=3, value=v.completion_count).number_format = INTEGER
        ws.cell(row=row, column=4, value=v.visit_total).number_format = MONEY
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BOX
        row += 1

    # Visit totals summary
    ws.cell(row=row, column=1, value="PROCEDURES SUBTOTAL")
    ws.cell(row=row, column=4, value=comp.procedures_subtotal).number_format = MONEY
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).font = TOTAL_FONT
        ws.cell(row=row, column=c).border = BOX
    row += 2

    # Site Fees block
    ws.cell(row=row, column=1, value="SITE FEES").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    fee_headers = ["Fee", "Sponsor Proposed", "Site Amount", "Frequency"]
    for c, h in enumerate(fee_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(fee_headers))
    row += 1

    for f in comp.site_fees:
        ws.cell(row=row, column=1, value=f.name + ("  *" if f.overridden else ""))
        if f.sponsor_proposed is not None:
            ws.cell(row=row, column=2, value=f.sponsor_proposed).number_format = MONEY
        else:
            ws.cell(row=row, column=2, value="Not Included")
        ws.cell(row=row, column=3, value=f.site_amount).number_format = MONEY
        ws.cell(row=row, column=4, value=f.frequency)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BOX
        row += 1

    ws.cell(row=row, column=1, value="SITE FEES SUBTOTAL")
    ws.cell(row=row, column=3, value=comp.site_fees_subtotal).number_format = MONEY
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).font = TOTAL_FONT
        ws.cell(row=row, column=c).border = BOX
    row += 2

    # Pass-through block
    ws.cell(row=row, column=1, value="PASS-THROUGH COSTS (Direct Reimbursement)").font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    for c, h in enumerate(fee_headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(fee_headers))
    row += 1

    for f in comp.pass_throughs:
        ws.cell(row=row, column=1, value=f.name + ("  *" if f.overridden else ""))
        ws.cell(
            row=row,
            column=2,
            value=f.sponsor_proposed if f.sponsor_proposed is not None else "Not Included",
        )
        if f.sponsor_proposed is not None:
            ws.cell(row=row, column=2).number_format = MONEY
        ws.cell(row=row, column=3, value=f.site_amount).number_format = MONEY
        ws.cell(row=row, column=4, value=f.frequency)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BOX
        row += 1

    ws.cell(row=row, column=1, value="PASS-THROUGH SUBTOTAL")
    ws.cell(row=row, column=3, value=comp.pass_throughs_subtotal).number_format = MONEY
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).font = TOTAL_FONT
        ws.cell(row=row, column=c).border = BOX
    row += 2

    # Grand totals
    summary_rows = [
        ("PROCEDURES (TRIAL TOTAL)", comp.procedures_subtotal),
        ("SITE FEES TOTAL", comp.site_fees_subtotal),
        ("PASS-THROUGH TOTAL", comp.pass_throughs_subtotal),
        ("GRAND TOTAL", comp.grand_total),
        ("  Sponsor portion", comp.sponsor_grand_total),
        ("  Medicare portion", comp.medicare_grand_total),
    ]
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=3, value=value).number_format = MONEY
        is_grand = "GRAND" in label
        for c in range(1, 5):
            cell = ws.cell(row=row, column=c)
            cell.border = BOX
            if is_grand:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
        row += 1

    _autosize(ws, {1: 60, 2: 22, 3: 22, 4: 30})

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- PRA workbook ---

def write_pra_workbook(comp: ComputedBudget, trial_name: str) -> bytes:
    """3-sheet PRA workbook mirroring the structure of the reference file:
    Coverage Analysis (per procedure), Research Prices (flat), Visit Cost Build.
    """
    # Group lines by procedure for the per-procedure sheets, by visit for the build.
    by_proc: dict[int, list] = defaultdict(list)
    by_visit: dict[str, list] = defaultdict(list)
    for line in comp.procedure_lines:
        by_proc[line.procedure_id].append(line)
        by_visit[line.visit_label].append(line)

    # First-occurrence price for each procedure (overrides are uniform per round
    # so any line for that proc has the same unit price).
    proc_summary = []
    for pid, lines in by_proc.items():
        sample = lines[0]
        applicable_visits = sorted({l.visit_label for l in lines})
        proc_summary.append({
            "procedure_id": pid,
            "code": sample.code,
            "name": sample.name,
            "category": sample.category,
            "coverage_status": sample.coverage_status,
            "unit_amc_base": sample.unit_amc_base,
            "unit_overhead_pct": sample.unit_overhead_pct,
            "unit_total_with_oh": sample.unit_total_with_oh,
            "applicable_visits": applicable_visits,
            "overridden": sample.overridden,
        })
    proc_summary.sort(key=lambda p: (p["category"] or "", p["code"]))

    wb = Workbook()

    # ----- Sheet 1: PRA-Coverage Analysis -----
    ws1 = wb.active
    ws1.title = "PRA-Coverage Analysis"
    ws1.cell(row=1, column=1, value=f"PRA — COVERAGE ANALYSIS — {trial_name}").font = Font(bold=True, size=14)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws1.cell(row=2, column=1, value=f"Round {comp.round_number}: {comp.round_label}{'  (FROZEN)' if comp.is_frozen else ''}").font = Font(italic=True)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = [
        "CPT / Code",
        "Procedure / Line Item",
        "Category",
        "Coverage Status",
        "AMC Base",
        "OH %",
        "AMC Total (incl. OH)",
        "Visits Applicable",
    ]
    for c, h in enumerate(headers, 1):
        ws1.cell(row=4, column=c, value=h)
    _style_header_row(ws1, 4, len(headers))

    row = 5
    current_category = None
    for p in proc_summary:
        if p["category"] != current_category:
            current_category = p["category"]
            ws1.cell(row=row, column=1, value=f"  ▶  {(current_category or 'Uncategorized').upper()}")
            ws1.cell(row=row, column=1).font = SECTION_FONT
            for c in range(1, len(headers) + 1):
                ws1.cell(row=row, column=c).fill = SECTION_FILL
            row += 1

        ws1.cell(row=row, column=1, value=p["code"])
        ws1.cell(row=row, column=2, value=p["name"] + ("  *" if p["overridden"] else ""))
        ws1.cell(row=row, column=3, value=p["category"])
        ws1.cell(row=row, column=4, value=p["coverage_status"].replace("_", " ").title())
        ws1.cell(row=row, column=5, value=p["unit_amc_base"]).number_format = MONEY
        ws1.cell(row=row, column=6, value=p["unit_overhead_pct"]).number_format = PERCENT
        ws1.cell(row=row, column=7, value=p["unit_total_with_oh"]).number_format = MONEY
        ws1.cell(row=row, column=8, value=", ".join(p["applicable_visits"]))
        for c in range(1, len(headers) + 1):
            ws1.cell(row=row, column=c).border = BOX
        row += 1

    _autosize(ws1, {1: 14, 2: 50, 3: 24, 4: 22, 5: 14, 6: 10, 7: 18, 8: 60})

    # ----- Sheet 2: PRA Research Prices -----
    ws2 = wb.create_sheet("PRA Research Prices")
    ws2.cell(row=1, column=1, value=f"PRA — RESEARCH PRICES — {trial_name}").font = Font(bold=True, size=14)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    price_headers = ["Code", "Procedure / Line Item", "AMC Base", "OH %", "AMC Total (incl. OH)"]
    for c, h in enumerate(price_headers, 1):
        ws2.cell(row=3, column=c, value=h)
    _style_header_row(ws2, 3, len(price_headers))

    row = 4
    current_category = None
    for p in proc_summary:
        if p["category"] != current_category:
            current_category = p["category"]
            ws2.cell(row=row, column=1, value=f"  ▶  {(current_category or 'Uncategorized').upper()}")
            ws2.cell(row=row, column=1).font = SECTION_FONT
            for c in range(1, len(price_headers) + 1):
                ws2.cell(row=row, column=c).fill = SECTION_FILL
            row += 1
        ws2.cell(row=row, column=1, value=p["code"])
        ws2.cell(row=row, column=2, value=p["name"])
        ws2.cell(row=row, column=3, value=p["unit_amc_base"]).number_format = MONEY
        ws2.cell(row=row, column=4, value=p["unit_overhead_pct"]).number_format = PERCENT
        ws2.cell(row=row, column=5, value=p["unit_total_with_oh"]).number_format = MONEY
        for c in range(1, len(price_headers) + 1):
            ws2.cell(row=row, column=c).border = BOX
        row += 1

    _autosize(ws2, {1: 14, 2: 50, 3: 14, 4: 10, 5: 20})

    # ----- Sheet 3: PRA-Visit Cost Build -----
    ws3 = wb.create_sheet("PRA-Visit Cost Build")
    ws3.cell(row=1, column=1, value=f"PER-VISIT COST BUILD — {trial_name}").font = Font(bold=True, size=14)
    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    build_headers = [
        "Visit",
        "Code",
        "Procedure",
        "Coverage",
        "Unit Total (incl. OH)",
        "Completion Count",
        "Sponsor Portion",
        "Medicare Portion",
    ]
    for c, h in enumerate(build_headers, 1):
        ws3.cell(row=3, column=c, value=h)
    _style_header_row(ws3, 3, len(build_headers))

    # Maintain visit order from comp.visits
    visit_order = [v.visit_label for v in comp.visits]
    visit_label_to_summary = {v.visit_label: v for v in comp.visits}

    row = 4
    for visit_label in visit_order:
        ws3.cell(row=row, column=1, value=f"  ▶  {visit_label}")
        ws3.cell(row=row, column=1).font = SECTION_FONT
        for c in range(1, len(build_headers) + 1):
            ws3.cell(row=row, column=c).fill = SECTION_FILL
        row += 1

        for line in by_visit[visit_label]:
            ws3.cell(row=row, column=2, value=line.code)
            ws3.cell(row=row, column=3, value=line.name + ("  *" if line.overridden else ""))
            ws3.cell(row=row, column=4, value=line.coverage_status.replace("_", " ").title())
            ws3.cell(row=row, column=5, value=line.unit_total_with_oh).number_format = MONEY
            ws3.cell(row=row, column=6, value=line.completion_count).number_format = INTEGER
            ws3.cell(row=row, column=7, value=line.sponsor_portion).number_format = MONEY
            ws3.cell(row=row, column=8, value=line.medicare_portion).number_format = MONEY
            for c in range(1, len(build_headers) + 1):
                ws3.cell(row=row, column=c).border = BOX
            row += 1

        # Visit total row
        v = visit_label_to_summary[visit_label]
        ws3.cell(row=row, column=1, value=f"  VISIT TOTAL: {visit_label}")
        ws3.cell(row=row, column=5, value=v.visit_total).number_format = MONEY
        ws3.cell(row=row, column=7, value=v.sponsor_total).number_format = MONEY
        ws3.cell(row=row, column=8, value=v.medicare_total).number_format = MONEY
        for c in range(1, len(build_headers) + 1):
            ws3.cell(row=row, column=c).fill = TOTAL_FILL
            ws3.cell(row=row, column=c).font = TOTAL_FONT
            ws3.cell(row=row, column=c).border = BOX
        row += 2

    # Grand total row
    ws3.cell(row=row, column=1, value="GRAND TOTAL")
    ws3.cell(row=row, column=5, value=comp.procedures_subtotal).number_format = MONEY
    ws3.cell(row=row, column=7, value=sum(v.sponsor_total for v in comp.visits)).number_format = MONEY
    ws3.cell(row=row, column=8, value=sum(v.medicare_total for v in comp.visits)).number_format = MONEY
    for c in range(1, len(build_headers) + 1):
        ws3.cell(row=row, column=c).fill = TOTAL_FILL
        ws3.cell(row=row, column=c).font = Font(bold=True, size=12)
        ws3.cell(row=row, column=c).border = BOX

    _autosize(ws3, {1: 30, 2: 14, 3: 50, 4: 22, 5: 18, 6: 16, 7: 16, 8: 16})

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
